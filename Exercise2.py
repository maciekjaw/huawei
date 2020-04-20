from openpyxl import load_workbook
import requests
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Data"
filename = 'Task2.xlsx'
wb.save(filename)

def get_values():

    wb = load_workbook(filename)
    sheet = wb['Data']
    id = 1000010
    id_row = 2
    sheet.cell(row=1, column=1, value="ID")
    sheet.cell(row=1, column=2, value="Author")
    sheet.cell(row=1, column=3, value="Title")
    column_number = 2
    bibs = 'bibs'
    author_key = 'author'
    title_key = 'title'
    id_key = 'id'

    for x in range(1,11):
        url = "http://data.bn.org.pl/api/bibs.json?limit=20&sinceId=" + str(id) + "&fbclid=IwAR0_RcYQvLWxvX0uLhW07fttFvzfhaNFpMJy_F5xIYnB7aPgL5xE0L34Qrs"
        response = requests.get(url)
        json = response.json()
        id = json[bibs][0][id_key]
        author = json[bibs][0][author_key]
        title = json[bibs][0][title_key]
        sheet.cell(id_row, column=1).value = id
        sheet.cell(id_row, column=2).value = author
        sheet.cell(id_row, column=3).value = title
        column = str(chr(64 + column_number))
        sheet.column_dimensions[column].width = 70
        wb.save(filename)
        id +=1
        id_row +=1
        column_number += 1

get_values()