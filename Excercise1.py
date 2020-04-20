from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Data"
filename = 'Task1.xlsx'
wb.save(filename)

path = "C:\\receivedData\\"
source = list(Path(path).rglob("*.[xX][lL][sS][xX]"))

def get_values():

    wb = load_workbook(filename)
    sheet = wb['Data']
    for x in source:
        wb = load_workbook(x)
        sheet1 = wb['Employee Data']
        for i in range(1, sheet.max_row + 1):
            j=1
            if sheet.cell(row=i, column=j).value is None:
                sheet.cell(row=i, column=j).value = sheet1.cell(row=i, column=j).value
                wb.save(filename)
            else:
                sheet.cell(row=i+1, column=j+1).value = sheet1.cell(row=i+1, column=j+1).value
                wb.save(filename)
                j+=1
                i+=1

get_values()


